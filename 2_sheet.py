from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()
ws.title = "MySheet" # sheet name chage
ws.sheet_properties.tabColor = "ff66ff"
ws1 = wb.create_sheet("YourSheet") #sheet created as given.
ws2 = wb.create_sheet("NewSheet", 2)

new_ws = wb["NewSheet"] # Dict form sheet access

print(wb.sheetnames)

# Sheet copy
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

print(wb.sheetnames)

wb.save("Sample.xlsx")
